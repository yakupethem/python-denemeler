from openpyxl import *

kitap = Workbook()
sheet1 = kitap.active
kitap.save("C:\\Users\yakup\Desktop\python\hafta_3\hafta_3\dosya.xlsx")

kitap.create_sheet("y40",1)
kitap.get_active_sheet()
#kitap.get_sheet_names()
#kitap = Workbook()
sheet1 = kitap.active
sheet1.append(("yazılım", "hanem", "", "sutun4"))
sheet1.append(["python", "excel", "yazısı"])

kitap.save("dosya.xlsx")
kitap.close()
kitap = load_workbook("C:\\Users\yakup\Desktop\python\hafta_3\hafta_3\dosya.xlsx")
sheet=kitap.create_sheet("calismasayfasi1",1)
sheet.cell(row=5,column=2,value="Yazılım")
kitap.save("dosya.xlsx")
kitap.close()
