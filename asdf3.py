from  openpyxl import *
kitap = load_workbook("C:\\Users\yakup\Desktop\python\hafta_3\hafta_3\dosya.xlsx")
sheet=kitap.create_sheet("calismasayfasi1",1)
sheet.cell(row=5,column=2,value="Yazılım")
kitap.save("dosya.xlsx")
kitap.close()
